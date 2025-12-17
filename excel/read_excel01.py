from openpyxl import Workbook, load_workbook
from pathlib import Path
current_dir =  Path(__file__).resolve().parent
wb = load_workbook(current_dir / "test.xlsx")
ws = wb.active
print(ws["A2"].value)
print(ws["B2"].value)
