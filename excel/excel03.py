from openpyxl import Workbook, load_workbook
from pathlib import Path
#print(openpyxl.__version__)
current_dir =  Path(__file__).resolve().parent

wb = load_workbook(current_dir / "score.xlsx", data_only=True)
ws = wb.active
for row in ws.iter_rows(values_only=True):
  print(row)