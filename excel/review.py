# pip install openpyxl
from openpyxl import load_workbook, Workbook
from pathlib import Path
current_dir = Path(__file__).parent
data_list=[
  {"name":"홍길동","age":30,"address":"일산"},
  {"name":"김철수","age":25,"address":"서울"},
  {"name":"최영희","age":35,"address":"파주"}
]
data_list02=[
  ("홍길동",30,"일산"),
  ("김철수",25,"서울"),
  ("최영희",35,"파주")
]
data_list03=[
  ["홍길동",30,"일산"],
  ["김철수",25,"서울"],
  ["최영희",35,"파주"]
] 


wb = Workbook()
ws = wb.active
ws.title = "학생정보"
# ws.append("이름","나이","주소"])
# for data in data_list:
#     ws.append([data["name"], data["age"], data["address"]])
# wb.save(current_dir / "students.xlsx")

# ws.append(("이름","나이","주소"))
# for data in data_list02:
#     ws.append(data)
ws.append(("이름","나이","주소"))
for data in data_list03:
  ws.append(data)
wb.save(current_dir / "students03.xlsx")

#리스트 튜플은 그대로 저장 가능
#딕셔너리는 키값을 이용해서 값만 추출 후 저장해야함

wb02 =  Workbook()
ws02 = wb02.active
ws02.title  = "수식"
ws02["A1"] = "값01"
ws02["B1"] = "값02"
ws02["C1"] = "합계"
ws02["A2"] = 100
ws02["B2"] = 200
ws02["C2"] = "=SUM(A2:B2)"
wb02.save(current_dir / "formula.xlsx")

wb03 =  Workbook()
ws03 = wb03.active
ws03.title  = "수식"
scores = [90, 85, 88, 92, 95]
ws03["A1"] = "점수"
for idx, score in enumerate(scores, start=2):
    ws03[f"A{idx}"] = score   
ws03["A7"] = "=SUM(A2:A6)"    
ws03["A8"] = "=AVERAGE(A2:A6)"
wb03.save(current_dir / "formula02.xlsx")

data04 = [
    ("홍길동", 80, 90, 70),
    ("김철수", 60, 65, 55),
    ("최영희", 88, 92, 91)
]
#이름,국어,영어,수학,총점,평균, 합격  (70이상이면 합격)
wb04 = Workbook()
ws04 = wb04.active  
ws04.title = "성적표"
ws04.append(("이름","국어","영어","수학","총점","평균","합격/불합격"))
for idx, (name, kor, eng, math) in enumerate(data04, start=2):
   ws04.append((name, kor, eng, math,"","",""))
   ws04[f"E{idx}"] = f"=SUM(B{idx}:D{idx})"
   ws04[f"F{idx}"] = f"=AVERAGE(B{idx}:D{idx})"
   ws04[f"G{idx}"] = f'=IF(F{idx}>=70,"합격","불합격")'
wb04.save(current_dir / "scorecard.xlsx")
    

