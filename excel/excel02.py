from openpyxl import Workbook
from pathlib import Path
#print(openpyxl.__version__)
current_dir =  Path(__file__).resolve().parent
wb = Workbook()
ws = wb.active
headers = ["이름","국어","영어","수학","평균","등급"]
ws.append(headers)
data_list = [
  ("홍길동",90,85,87),
  ("이순신",95,97,98),
  ("강감찬",86,50,92),
]
for row in data_list:
  ws.append(list(row)+["",""])

for r in range(2,2+len(data_list)):
  ws[f"E{r}"] = f"=AVERAGE(B{r}:D{r})"
  ws[f"F{r}"] = f'=IF(E{r}>=90,"A",IF(E{r}>=80,"B","C"))'

wb.save(current_dir / "score.xlsx")
