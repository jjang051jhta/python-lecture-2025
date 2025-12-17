# pip install openpyxl
from openpyxl import Workbook
from pathlib import Path
#print(openpyxl.__version__)
current_dir =  Path(__file__).resolve().parent
wb = Workbook()
ws = wb.active
# ws.title = "엑셀01"
# ws["A1"] = "이름"
# ws["B1"] = "점수"
# ws["A2"] = "홍길동"
# ws["B2"] = 90
# wb.save(current_dir / "test.xlsx")

data_list = [
  {"name":"홍길동","score":90},
  {"name":"이순신","score":95},
  {"name":"도요토미","score":50},
]
ws.append(["이름","점수"])
for data in data_list:
  ws.append([data["name"],data["score"]])
wb.save(current_dir / "test.xlsx")
