import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

def auto_width(ws):
  for col_idx in range(1,ws.max_column+1):
    col_letter = get_column_letter(col_idx) # 번호로 접급이 안되서 문자로 바꿔서 접근
    #print(col_letter)
    max_len = 0
    for cell in  ws[col_letter]:
      if cell.value is None:
        continue
      max_len = max(max_len,len(str(cell.value)))
    ws.column_dimensions[col_letter].width=max_len+4


current_dir = Path(__file__).resolve().parent

url = "https://webscraper.io/test-sites/e-commerce/allinone/computers/tablets"
res = requests.get(url)
res.raise_for_status()
soup = BeautifulSoup(res.text,"html.parser")
#print(soup.prettify())
product_list = soup.select(".product-wrapper")
#print(len(product_list))
data_list=[]
for product in product_list:
  title =  product.select_one("a.title").get_text(strip=True)
  price = product.select_one("h4.price").get_text(strip=True)
  link = "https://webscraper.io"+product.select_one("a").get("href","")
  data_list.append((title,price,link))

wb = Workbook()
ws = wb.active
ws.title="tablet-popular-10"
ws.append(["번호","상품명","가격","링크"])
items= []
for idx,(title,price,link)  in enumerate(data_list):
  print(f"{idx}/{title}")
  items.append([idx,title,price,link])



for item in items:
  ws.append(item)
#print(f"ws.max_column == {ws.max_column}")
# ws.column_dimensions["A"].width=10
# ws.column_dimensions["B"].width=30
# ws.column_dimensions["C"].width=15
# ws.column_dimensions["D"].width=40
auto_width(ws)
wb.save(current_dir / "tablet-popular-10.xlsx")

