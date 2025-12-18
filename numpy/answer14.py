import numpy as np
from pathlib import Path
from faker import Faker
from openpyxl import Workbook, load_workbook

"""
np.random.seed(7)
N = 10
faker = Faker("ko_KR")
faker.seed_instance(7)
# print(faker.name())
# print(faker.email())
# print(faker.job())
names = [faker.name() for _ in range(N)]
emails = [faker.email() for _ in range(N)]
scores = np.random.randint(50, 101, size=(N, 3))  # 50~100점 사이의 정수 난수 생성  
print(names)

total_scores = scores.sum(axis=1) # 각 학생의 총점 계산
avg_scores = scores.mean(axis=1)  # 각 학생의 평균 계산
#print(total_scores)
#print(avg_scores)
wb = Workbook()
ws = wb.active
ws.title = "학생성적"
headers = ["이름", "이메일", "국어", "영어", "수학", "총점", "평균"]
ws.append(headers)
for idx in range(N):
    row = [
        names[idx],
        emails[idx],
        scores[idx, 0],
        scores[idx, 1],
        scores[idx, 2],
        total_scores[idx],
        round(avg_scores[idx], 2)
    ]
    ws.append(row)  
file_path = Path(__file__).parent / "student_scores.xlsx"    
wb.save(file_path)
"""
# current_dir = Path(__file__).resolve().parent
# wb = load_workbook(current_dir / "student_scores.xlsx")
# ws =  wb.active
# rows = []

# for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
#     rows.append(list(row))

# for row in range(2, ws.max_row + 1):
#     rows.append([
#             ws[f"A{row}"].value,
#             ws[f"B{row}"].value,  
#             ws[f"C{row}"].value,
#             ws[f"D{row}"].value,
#             ws[f"E{row}"].value,
#             ws[f"F{row}"].value,
#             ws[f"G{row}"].value
#     ])
#print(rows)    
# data = np.array(rows, dtype=object)  # object로 해야 숫자도 같이 들어감
# print(data)
# avg_scores = data[:, 6].astype(int)  # 국어, 영어, 수학 점수 부분만 추출하여 정수형으로 변환
# print(avg_scores)
# mask =  avg_scores >= 70
# passed_students = data[mask]
# print(passed_students)
# wb = Workbook()
# ws = wb.active
# ws.title = "성적70점이상"
# headers = ["이름", "이메일", "국어", "영어", "수학", "총점", "평균"]
# ws.append(headers)
# for row in passed_students:
#     ws.append(row.tolist())
# file_path = Path(__file__).parent / "student_passed.xlsx"    
# wb.save(file_path)
    

# current_dir = Path(__file__).resolve().parent
# wb = load_workbook(current_dir / "student_scores.xlsx")
# ws =  wb.active
# scores=[]
# # 엑셀의 상단 header의 A/B/C...열을 기준으로 값을 읽을 때는 문자열로 접근
# for row in range(2, ws.max_row+1):
#     scores.append([
#                    ws.cell(row,1).value, 
#                    ws.cell(row,2).value, 
#                    ws.cell(row,3).value, 
#                    ws.cell(row,4).value, 
#                    ws.cell(row,5).value, 
#                    ws.cell(row,6).value, 
#                    ws.cell(row,7).value])
# print(scores)
# data = np.array(scores, dtype=object)
# subjects = data[:, 2:5].astype(int)
# print(subjects)
# subjects_mean = subjects.mean(axis=0)
# subjects_max = subjects.max(axis=0)
# subjects_min = subjects.min(axis=0)
# subjects_std = subjects.std(axis=0)
# print(f"과목별 평균 : {subjects_mean}")
# print(f"과목별 최대값 : {subjects_max}")    
# print(f"과목별 최소값 : {subjects_min}")
# print(f"과목별 표준편차 : {subjects_std}")

# 표준편차가 크다 => 점수가 흩어져 있다. (평균에 모여있지 않다.)  
# 표준편차가 작다 => 점수가 모여 있다.  (평균에 수렴한다.)

# test_data = np.array([10,10,10,10,10,9,9,9,9])
# print(np.std(test_data))
# test_data02 = np.array([100,90,10,10,10,9,9,9,80])
# print(np.std(test_data02))
import numpy as np
from pathlib import Path
from faker import Faker
from openpyxl import Workbook

# np.random.seed(7)
# N = 10
# faker = Faker("ko_KR")
# faker.seed_instance(7)
# print(faker.name())
# print(faker.email())
# print(faker.job())
# names = [faker.name() for _ in range(N)]
# emails = [faker.email() for _ in range(N)]
# scores = np.random.randint(50, 101, size=(N, 3))  # 50~100점 사이의 정수 난수 생성  
# print(names)

# total_scores = scores.sum(axis=1) # 각 학생의 총점 계산
# avg_scores = scores.mean(axis=1)  # 각 학생의 평균 계산
# grades= np.where(avg_scores >= 90, "A",
#                   np.where(avg_scores >= 80, "B",
#                     np.where(avg_scores >= 70, "C",
#                       np.where(avg_scores >= 60, "D", "F"))))
         
#print(total_scores)
#print(avg_scores)
# wb = Workbook()
# ws = wb.active
# ws.title = "학생성적"
# headers = ["이름", "이메일", "국어", "영어", "수학", "총점", "평균", "학점"]
# ws.append(headers)
# for idx in range(N):
#     row = [
#         names[idx],
#         emails[idx],
#         scores[idx, 0],
#         scores[idx, 1],
#         scores[idx, 2],
#         total_scores[idx],
#         round(avg_scores[idx], 2),
#         grades[idx]
#     ]
#     ws.append(row)  
# file_path = Path(__file__).parent / "student_scores.xlsx"    #이건 일반 .py에서는 이렇게 쓰면 된다.
# #file_path = Path.cwd() / "student_scores.xlsx"  #이건 Jupyter Notebook에서는 이렇게 써야 한다.b.save(file_path)
# wb.save(file_path)




np.random.seed(7)
N = 20
faker = Faker("ko_KR")
faker.seed_instance(7)
names = [faker.name() for _ in range(N)]
salaries = np.random.randint(2500000, 5000001, size=N)  # 300만원~800만원 사이의 정수 난수 생성
departments = np.random.choice(["개발부", "영업부", "인사부", "총무부"], size=N)
unique_departments = np.unique(departments)
#print(departments)
#print(unique_departments)
#print(names)
#print(salaries)

dept_avg_salaries = {}
for dept in unique_departments:
    mask = departments == dept
    dept_salaries = salaries[mask].mean()
    dept_avg_salaries[dept] =   int(dept_salaries)  

#print(dept_avg_salaries)
for dept, avg_salary in dept_avg_salaries.items():
    print(f"{dept} : {avg_salary}원")


wb = Workbook()
ws = wb.active
ws.title = "직원급여"
headers = ["이름", "부서","급여"]
for idx in range(N):
    row = [
        names[idx],
        departments[idx],
        salaries[idx],
    ]
    ws.append(row)
ws02 = wb.create_sheet(title="부서별평균급여")
ws02.append(["부서", "평균급여"])      
for dept, avg_salary in dept_avg_salaries.items():
    ws02.append([dept, avg_salary])

file_path = Path(__file__).parent / "employee_salaries.xlsx"  
wb.save(file_path)